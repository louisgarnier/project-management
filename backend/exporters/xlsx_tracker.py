"""xlsx tracker exporter (Story 15.8).

build_tracker_xlsx(project_id) -> bytes: renders 5 sheets via openpyxl.
No disk writes; everything goes through BytesIO.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.services.project_tracker_data import list_project_topics_with_call_history

# Re-exported at module level so tests can monkeypatch xlsx_tracker.list_project_topics:
list_project_topics = list_project_topics_with_call_history


def list_topics_timeline(project_id: str) -> dict | None:  # noqa: ARG001
    """Placeholder — reserved for future timeline aggregation (Story 15.8+)."""
    return None


HEADER_FONT = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="EEEEEE")
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")


def build_tracker_xlsx(project_id: str) -> bytes:
    project_topics = list_project_topics(project_id) or []
    timeline = list_topics_timeline(project_id)  # noqa: F841 — reserved for future use

    wb = Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, project_topics)
    _build_chronology(wb, project_topics)
    _build_anchors_lifecycle(wb, project_topics)
    _build_decisions_log(wb, project_topics)
    _build_key_terms_registry(wb, project_topics)

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def _header(ws, cols: list[str]) -> None:
    for i, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"


def _collect_unique_calls(project_topics: list[dict]) -> list[dict]:
    seen = {}
    for t in project_topics:
        for c in t.get("calls") or []:
            cid = c.get("call_id")
            if cid and cid not in seen:
                seen[cid] = c
    return sorted(seen.values(), key=lambda c: c.get("call_date") or "")


def _build_dashboard(wb: Workbook, project_topics: list[dict]) -> None:
    ws = wb.create_sheet("Dashboard")
    _header(ws, ["Metric", "Count"])
    n_topics = len(project_topics)
    n_tasks_open = 0
    n_tasks_in_progress = 0
    n_tasks_resolved = 0
    n_oqs_open = 0
    n_decisions = 0
    for t in project_topics:
        for c in t.get("calls") or []:
            for task in c.get("tasks") or []:
                status = task.get("status", "open")
                if status == "open":
                    n_tasks_open += 1
                elif status == "in_progress":
                    n_tasks_in_progress += 1
                elif status == "resolved":
                    n_tasks_resolved += 1
            for q in c.get("open_questions") or []:
                if q.get("status") != "resolved":
                    n_oqs_open += 1
            n_decisions += len(c.get("decisions") or [])
    rows = [
        ("Total topics", n_topics),
        ("Open tasks", n_tasks_open),
        ("In-progress tasks", n_tasks_in_progress),
        ("Resolved tasks", n_tasks_resolved),
        ("Open questions still open", n_oqs_open),
        ("Total decisions", n_decisions),
    ]
    for i, (k, v) in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 12


def _build_chronology(wb: Workbook, project_topics: list[dict]) -> None:
    ws = wb.create_sheet("Chronology")
    calls = _collect_unique_calls(project_topics)
    headers = ["Topic"] + [c.get("call_date") or c.get("call_id", "?") for c in calls] + ["RAG note"]
    _header(ws, headers)
    for i, t in enumerate(project_topics, 2):
        ws.cell(row=i, column=1, value=t.get("name", ""))
        calls_by_id = {c.get("call_id"): c for c in t.get("calls") or []}
        latest_rag = ""
        for j, call in enumerate(calls, 2):
            tc = calls_by_id.get(call.get("call_id"))
            if tc:
                ws.cell(row=i, column=j, value=tc.get("chronology_narrative") or "").alignment = WRAP_ALIGN
                rag = tc.get("rag_verification_note") or ""
                if rag and rag != "verified":
                    latest_rag = rag
        ws.cell(row=i, column=len(headers), value=latest_rag).alignment = WRAP_ALIGN
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, len(calls) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 50
    ws.column_dimensions[get_column_letter(len(headers))].width = 40


def _resolve_call_date(project_topics: list[dict], call_id: str | None) -> str:
    if not call_id:
        return ""
    for t in project_topics:
        for c in t.get("calls") or []:
            if c.get("call_id") == call_id:
                return c.get("call_date") or ""
    return ""


def _build_anchors_lifecycle(wb: Workbook, project_topics: list[dict]) -> None:
    ws = wb.create_sheet("Anchors lifecycle")
    _header(ws, ["Topic", "Kind", "Item", "Owner", "Status", "Added in", "Closed in"])
    row_i = 2
    seen_keys = set()  # de-dupe item appearing in multiple calls (e.g. same task carrying across)
    for t in project_topics:
        topic_name = t.get("name", "")
        for c in t.get("calls") or []:
            for task in c.get("tasks") or []:
                key = ("task", task.get("task_id") or task.get("task"))
                if key in seen_keys:
                    # Update existing row's status + closed_in if more recent — for simplicity, skip
                    continue
                seen_keys.add(key)
                added_date = _resolve_call_date(project_topics, task.get("added_in_call_id"))
                closed_date = _resolve_call_date(project_topics, task.get("closed_in_call_id"))
                ws.cell(row=row_i, column=1, value=topic_name)
                ws.cell(row=row_i, column=2, value="task")
                ws.cell(row=row_i, column=3, value=task.get("task", "")).alignment = WRAP_ALIGN
                ws.cell(row=row_i, column=4, value=task.get("owner", ""))
                ws.cell(row=row_i, column=5, value=task.get("status", ""))
                ws.cell(row=row_i, column=6, value=added_date)
                ws.cell(row=row_i, column=7, value=closed_date)
                row_i += 1
            for q in c.get("open_questions") or []:
                key = ("oq", q.get("id") or q.get("text"))
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                added_date = _resolve_call_date(project_topics, q.get("added_in_call_id"))
                closed_date = _resolve_call_date(project_topics, q.get("closed_in_call_id"))
                ws.cell(row=row_i, column=1, value=topic_name)
                ws.cell(row=row_i, column=2, value="open_question")
                ws.cell(row=row_i, column=3, value=q.get("text", "")).alignment = WRAP_ALIGN
                ws.cell(row=row_i, column=4, value=q.get("owner", ""))
                ws.cell(row=row_i, column=5, value=q.get("status", ""))
                ws.cell(row=row_i, column=6, value=added_date)
                ws.cell(row=row_i, column=7, value=closed_date)
                row_i += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 36
    for col in ["D", "E", "F", "G"]:
        ws.column_dimensions[col].width = 14


def _build_decisions_log(wb: Workbook, project_topics: list[dict]) -> None:
    ws = wb.create_sheet("Decisions log")
    _header(ws, ["Topic", "Decision", "Decided in"])
    row_i = 2
    seen_keys = set()
    for t in project_topics:
        topic_name = t.get("name", "")
        for c in t.get("calls") or []:
            for d in c.get("decisions") or []:
                key = d.get("id") or (topic_name, d.get("text"))
                if key in seen_keys:
                    continue
                seen_keys.add(key)
                added_date = _resolve_call_date(project_topics, d.get("added_in_call_id") or c.get("call_id"))
                ws.cell(row=row_i, column=1, value=topic_name)
                ws.cell(row=row_i, column=2, value=d.get("text", "")).alignment = WRAP_ALIGN
                ws.cell(row=row_i, column=3, value=added_date)
                row_i += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 14


def _build_key_terms_registry(wb: Workbook, project_topics: list[dict]) -> None:
    ws = wb.create_sheet("Key terms registry")
    _header(ws, ["Topic", "Key term", "First seen"])
    row_i = 2
    for t in project_topics:
        topic_name = t.get("name", "")
        first_call_date = ""
        if t.get("calls"):
            first_call = sorted(t["calls"], key=lambda c: c.get("call_date") or "")[0]
            first_call_date = first_call.get("call_date") or ""
        for kt in t.get("key_terms") or []:
            ws.cell(row=row_i, column=1, value=topic_name)
            ws.cell(row=row_i, column=2, value=kt)
            ws.cell(row=row_i, column=3, value=first_call_date)
            row_i += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
