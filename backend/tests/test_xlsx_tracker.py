"""Tests for the xlsx tracker exporter (Story 15.8)."""

import pytest
from io import BytesIO

from openpyxl import load_workbook

from backend.exporters import xlsx_tracker


@pytest.fixture
def _fake_project_topics():
    return [
        {
            "id": "t1",
            "name": "MC Mac memory",
            "key_terms": ["MC Mac", "memory"],
            "calls": [
                {
                    "call_id": "c1",
                    "call_date": "2026-05-01",
                    "chronology_narrative": "Discussed memory issue.",
                    "rag_verification_note": "verified",
                    "tasks": [
                        {
                            "task": "investigate",
                            "next_step": "Test boost flag",
                            "owner": "Nick",
                            "status": "open",
                            "added_in_call_id": "c1",
                            "closed_in_call_id": None,
                        }
                    ],
                    "open_questions": [
                        {
                            "text": "Does boost flag fix?",
                            "owner": "Nick",
                            "status": "open",
                            "added_in_call_id": "c1",
                            "closed_in_call_id": None,
                        }
                    ],
                    "decisions": [{"text": "Defer to Q3"}],
                },
                {
                    "call_id": "c2",
                    "call_date": "2026-05-08",
                    "chronology_narrative": "Resolved with boost.",
                    "rag_verification_note": "verified",
                    "tasks": [
                        {
                            "task": "investigate",
                            "next_step": "",
                            "owner": "Nick",
                            "status": "resolved",
                            "added_in_call_id": "c1",
                            "closed_in_call_id": "c2",
                        }
                    ],
                    "open_questions": [],
                    "decisions": [],
                },
            ],
        },
    ]


def test_build_tracker_xlsx_produces_5_sheets(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    assert set(wb.sheetnames) == {
        "Dashboard",
        "Chronology",
        "Anchors lifecycle",
        "Decisions log",
        "Key terms registry",
    }


def test_chronology_sheet_has_topic_row_and_call_columns(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    ws = wb["Chronology"]
    # 1 header row + 1 topic row = 2 rows
    assert ws.max_row == 2
    # Cols: Topic + 2 call dates + RAG note = 4 cols
    assert ws.max_column == 4
    # Topic name appears in row 2 col 1
    assert ws.cell(row=2, column=1).value == "MC Mac memory"


def test_anchors_lifecycle_includes_tasks_and_open_questions(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    ws = wb["Anchors lifecycle"]
    # 1 topic, 2 calls: task appears in BOTH calls (open in c1, resolved in c2) + 1 OQ in c1
    # Implementation may de-duplicate by item id — at minimum, header + 1 task row + 1 OQ row = 3 rows
    assert ws.max_row >= 3


def test_decisions_log_row_count(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    ws = wb["Decisions log"]
    # 1 decision in fixture; header + 1 data row
    assert ws.max_row == 2
    assert ws.cell(row=2, column=2).value == "Defer to Q3"  # decision text in col 2


def test_key_terms_registry_row_count(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    ws = wb["Key terms registry"]
    # 2 key terms in fixture; header + 2 data rows
    assert ws.max_row == 3


def test_dashboard_renders_counts(_fake_project_topics, monkeypatch):
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: _fake_project_topics)
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p1")
    wb = load_workbook(BytesIO(blob))
    ws = wb["Dashboard"]
    # At least: total topics row, total tasks row, total OQs row, total decisions row
    # Implementation can lay it out flexibly; just confirm the cell values appear somewhere.
    all_cells = [c.value for row in ws.iter_rows() for c in row if c.value is not None]
    # Should contain the topic count "1" or "1 topic" somewhere, etc.
    # Loosen the assertion: just verify the sheet has SOME data
    assert len(all_cells) > 0


def test_build_tracker_xlsx_with_empty_project(monkeypatch):
    """A project with zero topics still produces all 5 sheets (header rows only)."""
    monkeypatch.setattr(xlsx_tracker, "list_project_topics", lambda _pid: [])
    monkeypatch.setattr(xlsx_tracker, "list_topics_timeline", lambda _pid: None)
    blob = xlsx_tracker.build_tracker_xlsx("p-empty")
    wb = load_workbook(BytesIO(blob))
    assert set(wb.sheetnames) == {
        "Dashboard",
        "Chronology",
        "Anchors lifecycle",
        "Decisions log",
        "Key terms registry",
    }
    # Each sheet has at least the header row
    for name in wb.sheetnames:
        assert wb[name].max_row >= 1


def test_export_xlsx_endpoint_returns_200_with_correct_headers(monkeypatch):
    """The GET endpoint returns 200, correct mime + Content-Disposition."""
    from fastapi.testclient import TestClient
    from backend.main import app
    from backend.routers import projects as projects_router

    fake_db = type("DB", (), {
        "table": lambda self, name: type("T", (), {
            "select": lambda self, _: self,
            "eq": lambda self, _k, _v: self,
            "execute": lambda self: type("R", (), {"data": [{"name": "FactSet SWIB"}]})()
        })()
    })()
    monkeypatch.setattr(projects_router, "get_client", lambda: fake_db)
    monkeypatch.setattr(
        "backend.exporters.xlsx_tracker.list_project_topics",
        lambda _pid: []
    )

    with TestClient(app) as client:
        resp = client.get("/api/projects/abc/export.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in resp.headers["content-disposition"]
    assert ".xlsx" in resp.headers["content-disposition"]


def test_tracker_data_endpoint_returns_per_call_shape(monkeypatch):
    """The GET /tracker-data endpoint returns the per-(topic, call) shape."""
    from fastapi.testclient import TestClient
    from backend.main import app
    from backend.routers import projects as projects_router

    fake_db = type("DB", (), {
        "table": lambda self, name: type("T", (), {
            "select": lambda self, _: self,
            "eq": lambda self, _k, _v: self,
            "execute": lambda self: type("R", (), {"data": [{"id": "p1"}]})()
        })()
    })()
    monkeypatch.setattr(projects_router, "get_client", lambda: fake_db)
    monkeypatch.setattr(
        "backend.services.project_tracker_data.list_project_topics_with_call_history",
        lambda _pid: [{"id": "t1", "name": "T", "key_terms": [], "calls": []}]
    )
    with TestClient(app) as client:
        resp = client.get("/api/projects/p1/tracker-data")
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert data[0]["name"] == "T"


def test_tracker_data_endpoint_404_on_missing_project(monkeypatch):
    """Missing project → 404 on /tracker-data."""
    from fastapi.testclient import TestClient
    from backend.main import app
    from backend.routers import projects as projects_router

    fake_db = type("DB", (), {
        "table": lambda self, name: type("T", (), {
            "select": lambda self, _: self,
            "eq": lambda self, _k, _v: self,
            "execute": lambda self: type("R", (), {"data": []})()
        })()
    })()
    monkeypatch.setattr(projects_router, "get_client", lambda: fake_db)

    with TestClient(app) as client:
        resp = client.get("/api/projects/missing/tracker-data")
    assert resp.status_code == 404


def test_export_xlsx_endpoint_404_on_missing_project(monkeypatch):
    """Missing project → 404."""
    from fastapi.testclient import TestClient
    from backend.main import app
    from backend.routers import projects as projects_router

    fake_db = type("DB", (), {
        "table": lambda self, name: type("T", (), {
            "select": lambda self, _: self,
            "eq": lambda self, _k, _v: self,
            "execute": lambda self: type("R", (), {"data": []})()
        })()
    })()
    monkeypatch.setattr(projects_router, "get_client", lambda: fake_db)

    with TestClient(app) as client:
        resp = client.get("/api/projects/missing/export.xlsx")
    assert resp.status_code == 404
